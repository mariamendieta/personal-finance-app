"""
Portfolio analysis and Excel report generator.
Reads holdings from multiple account CSVs, normalizes them,
and produces a comprehensive portfolio report.
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

PORTFOLIO_ROOT = Path(__file__).parent / "InvestmentPortfolio"
# DATA_DIR and OUTPUT are set dynamically in __main__ based on snapshot date
DATA_DIR = None
OUTPUT = None

MONEY_FMT = '#,##0.00'
PCT_FMT = '0.0%'

# ── Styles ────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
GAIN_FONT = Font(color="006100")
LOSS_FONT = Font(color="9C0006")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

# ── Asset class mapping for known tickers/funds ──────────────────────────
# Used for files that don't include asset class info
ASSET_CLASS_MAP = {
    # Equity - US
    "VOO": "US Equity", "FXAIX": "US Equity", "VFIAX": "US Equity",
    "SFY": "US Equity", "IVE": "US Equity", "IVW": "US Equity",
    "DYNF": "US Equity", "THRO": "US Equity", "AMZN": "US Equity",
    "NVDA": "US Equity", "AAPL": "US Equity", "MSFT": "US Equity",
    "GOOGL": "US Equity", "META": "US Equity", "TSLA": "US Equity",
    "VGT": "US Equity", "XLF": "US Equity", "XLC": "US Equity",
    "XLY": "US Equity", "XLV": "US Equity", "XLI": "US Equity",
    "XLE": "US Equity", "XLP": "US Equity", "AVUV": "US Equity",
    "JTEK": "US Equity", "ARCC": "US Equity", "MAIN": "US Equity",
    "BAI": "US Equity", "SHLD": "US Equity", "CRM": "US Equity",
    "VTSAX": "US Equity",
    # Equity - International
    "VEA": "International Equity", "VWO": "International Equity",
    "VXUS": "International Equity", "FLIN": "International Equity",
    "VTMGX": "International Equity", "VEMAX": "International Equity",
    # Fixed Income
    "BILS": "Fixed Income", "JGCGX": "Fixed Income", "CBTAX": "Fixed Income",
    "VWIUX": "Fixed Income", "BRHYX": "Fixed Income", "VGIT": "Fixed Income",
    "VTEB": "Fixed Income", "SCHR": "Fixed Income", "SCHO": "Fixed Income",
    "BNDX": "International Fixed Income", "HIMU": "Fixed Income",
    "MUB": "Fixed Income", "PZA": "Fixed Income", "SPTL": "Fixed Income",
    "VBTLX": "Fixed Income", "VTABX": "International Fixed Income",
    # Alternatives
    "DLR": "Real Estate", "VNQ": "Real Estate", "VGSLX": "Real Estate",
    "MLPX": "Alternatives", "IAU": "Commodities (Gold)",
    "FETH": "Crypto", "ETHE": "Crypto",
    # Cash
    "MJLXX": "Cash", "QAJDS": "Cash", "QCERQ": "Cash", "VMFXX": "Cash",
    # Viviana's Roth
    "BND": "Fixed Income", "VBTLX": "Fixed Income", "AVDV": "International Equity", "SOXX": "US Equity",
}

# Map for Fidelity/401k fund names
FUND_ASSET_CLASS = {
    "Fidelity Total International Index": "International Equity",
    "Fidelity 500 Index": "US Equity",
    "Fidelity Mid Cap Index": "US Equity",
    "Fidelity U.S. Sustainability Index Instl": "US Equity",
    "BlackRock Balanced K": "Balanced/Multi-Asset",
    "DFA Intl Sustainability Core 1": "International Equity",
    "Fidelity Small Cap Index": "US Equity",
    "Hartford Dividend and Growth R6": "US Equity",
    "Undiscovered Mgrs Behavioral Value R6": "US Equity",
    "Putnam Small Cap Growth R6": "US Equity",
    "JPMorgan Large Cap Growth R6": "US Equity",
    "Invesco Discovery Mid Cap Growth R6": "US Equity",
    "PIMCO Income Instl": "Fixed Income",
    "Fidelity US Bond Index": "Fixed Income",
}

# Account type classification
ACCOUNT_TYPES = {
    "Chase Taxable Brokerage": "Taxable",
    "Chase Parametric": "Taxable",
    "Michael's Roth IRA (prod)": "Roth IRA",
    "Michael's Traditional IRA": "Traditional IRA",
    "SoFi Joint": "Taxable",
    "SoFi Self-Directed": "Taxable",
    "Viviana 401k": "401k",
    "Viviana Secondary 401k": "401k",
    "Viviana's Roth IRA (prod)": "Roth IRA",
    # Demo account names (no-op in production)
    "Joint Brokerage": "Taxable",
    "Michael's Roth IRA": "Roth IRA",
    "Michael's 401k": "401k",
    "Viviana's Roth IRA": "Roth IRA",
    "Viviana's 401k": "401k",
}


def parse_value(val):
    """Parse a value string like '$1,234.56' or '1,234.56' to float."""
    if pd.isna(val) or val == "-" or val == "–":
        return 0.0
    s = str(val).replace("$", "").replace(",", "").replace("+", "").strip()
    if s == "" or s == "-" or s == "–":
        return 0.0
    return float(s)


def parse_chase_holdings(filepath: Path, account_name: str) -> pd.DataFrame:
    """Parse Chase/JPMorgan CSV export (82-column format)."""
    df = pd.read_csv(filepath, on_bad_lines="skip")
    if "Asset Class" not in df.columns:
        return pd.DataFrame()

    # Remove footnotes
    mask = df["Asset Class"] == "FOOTNOTES"
    if mask.any():
        df = df.loc[: mask.idxmax() - 1]

    # Remove empty rows
    df = df[df["Asset Class"].notna() & (df["Asset Class"] != "")].copy()

    rows = []
    for _, r in df.iterrows():
        ticker = str(r.get("Ticker", "")).strip()
        desc = str(r.get("Description", "")).strip()
        asset_class_raw = str(r.get("Asset Class", "")).strip()
        value = parse_value(r.get("Value", 0))
        cost = parse_value(r.get("Cost", 0))
        qty = parse_value(r.get("Quantity", 0))
        price = parse_value(r.get("Price", 0))
        gl_pct = parse_value(r.get("Unrealized Gain/Loss (%)", 0))

        # Map asset class
        if ticker in ASSET_CLASS_MAP:
            asset_class = ASSET_CLASS_MAP[ticker]
        elif "Cash" in asset_class_raw or "Short Term" in str(r.get("Asset Strategy", "")):
            asset_class = "Cash"
        elif "Fixed Income" in asset_class_raw:
            asset_class = "Fixed Income"
        elif "Alternative" in asset_class_raw:
            strategy = str(r.get("Asset Strategy", ""))
            if "Real Estate" in strategy:
                asset_class = "Real Estate"
            else:
                asset_class = "Alternatives"
        elif "Equity" in asset_class_raw:
            strategy = str(r.get("Asset Strategy", ""))
            if "Global" in strategy or "Intl" in strategy:
                asset_class = "International Equity"
            elif "Smid" in strategy or "Small" in strategy:
                asset_class = "US Equity"
            else:
                asset_class = "US Equity"
        else:
            asset_class = "Other"

        rows.append({
            "account": account_name,
            "account_type": ACCOUNT_TYPES.get(account_name, "Unknown"),
            "ticker": ticker,
            "description": desc,
            "asset_class": asset_class,
            "quantity": qty,
            "price": price,
            "value": value,
            "cost": cost,
            "unrealized_gl": value - cost,
            "unrealized_gl_pct": gl_pct / 100 if gl_pct else (
                (value - cost) / cost if cost > 0 else 0
            ),
        })
    return pd.DataFrame(rows)


def parse_sofi_managed(filepath: Path, account_name: str) -> pd.DataFrame:
    """Parse SoFi managed portfolio (Ticker, Company, Current Allocation, Market Price, Unrealized Gain, Total Value)."""
    df = pd.read_csv(filepath)
    rows = []
    for _, r in df.iterrows():
        ticker = str(r.get("Ticker", "")).strip()
        desc = str(r.get("Company", "")).strip()
        value = parse_value(r.get("Total Value", 0))
        gl = parse_value(r.get("Unrealized Gain", 0))
        price = parse_value(r.get("Market Price", 0))
        cost = value - gl if gl != 0 else value

        if ticker and ticker != "nan" and ticker in ASSET_CLASS_MAP:
            asset_class = ASSET_CLASS_MAP[ticker]
        elif desc == "Cash" or (not ticker or ticker == "nan"):
            asset_class = "Cash"
        else:
            asset_class = "US Equity"

        rows.append({
            "account": account_name,
            "account_type": ACCOUNT_TYPES.get(account_name, "Taxable"),
            "ticker": ticker if ticker and ticker != "nan" else "CASH",
            "description": desc if desc and desc != "nan" else ticker,
            "asset_class": asset_class,
            "quantity": 0,
            "price": price,
            "value": value,
            "cost": cost,
            "unrealized_gl": gl,
            "unrealized_gl_pct": gl / cost if cost > 0 else 0,
        })
    return pd.DataFrame(rows)


def parse_sofi_brokerage(filepath: Path, account_name: str) -> pd.DataFrame:
    """Parse SoFi brokerage (Ticker, Shares, Market Price, Unrealized Gain, Total Value)."""
    df = pd.read_csv(filepath)
    rows = []
    for _, r in df.iterrows():
        ticker = str(r.get("Ticker", "")).strip()
        value = parse_value(r.get("Total Value", 0))
        gl = parse_value(r.get("Unrealized Gain", 0))
        price = parse_value(r.get("Market Price", 0))
        shares = parse_value(r.get("Shares", 0))
        cost = value - gl

        rows.append({
            "account": account_name,
            "account_type": ACCOUNT_TYPES.get(account_name, "Taxable"),
            "ticker": ticker,
            "description": ticker,
            "asset_class": ASSET_CLASS_MAP.get(ticker, "US Equity"),
            "quantity": shares,
            "price": price,
            "value": value,
            "cost": cost,
            "unrealized_gl": gl,
            "unrealized_gl_pct": gl / cost if cost > 0 else 0,
        })
    return pd.DataFrame(rows)



def parse_marias_roth(filepath: Path) -> pd.DataFrame:
    """Parse Viviana's Roth IRA (Symbol, Name, Price, Quantity, Current Balance)."""
    df = pd.read_csv(filepath, on_bad_lines="skip")
    rows = []
    for _, r in df.iterrows():
        ticker = str(r.get("Symbol", "")).strip()
        if ticker == "Total" or not ticker or ticker == "nan":
            continue
        desc = str(r.get("Name", "")).strip()
        value = parse_value(r.get("Current Balance", 0))
        gl = parse_value(r.get("$ Unrealized Gain/Loss", 0))
        price = parse_value(r.get("Price", 0))
        qty = parse_value(r.get("Quantity", 0))
        cost = value - gl if gl != 0 else value

        rows.append({
            "account": "Viviana's Roth IRA",
            "account_type": "Roth IRA",
            "ticker": ticker,
            "description": desc,
            "asset_class": ASSET_CLASS_MAP.get(ticker, "US Equity"),
            "quantity": qty,
            "price": price,
            "value": value,
            "cost": cost,
            "unrealized_gl": gl,
            "unrealized_gl_pct": gl / cost if cost > 0 else 0,
        })
    return pd.DataFrame(rows)


def parse_401k_simple(filepath: Path, account_name: str) -> pd.DataFrame:
    """Parse 401k with simple Investment/Balance format."""
    df = pd.read_csv(filepath)
    rows = []
    for _, r in df.iterrows():
        fund = str(r.get("Investment", "")).strip()
        value = parse_value(r.get("Balance", 0))
        asset_class = FUND_ASSET_CLASS.get(fund, "Other")

        rows.append({
            "account": account_name,
            "account_type": "401k",
            "ticker": "",
            "description": fund,
            "asset_class": asset_class,
            "quantity": 0,
            "price": 0,
            "value": value,
            "cost": 0,
            "unrealized_gl": 0,
            "unrealized_gl_pct": 0,
        })
    return pd.DataFrame(rows)


def parse_401k_vanguard(filepath: Path, account_name: str) -> pd.DataFrame:
    """Parse 401k with Vanguard format (Fund, Last 7 Days, Your All Time, Current Weight, # of Shares, Market Value)."""
    df = pd.read_csv(filepath)
    rows = []
    for _, r in df.iterrows():
        ticker = str(r.get("Fund", "")).strip()
        value = parse_value(r.get("Market Value", 0))
        shares = parse_value(r.get("# of Shares", 0))
        all_time_str = str(r.get("Your All Time", "0%")).replace("%", "").replace("+", "")
        try:
            all_time_pct = float(all_time_str) / 100
        except ValueError:
            all_time_pct = 0
        cost = value / (1 + all_time_pct) if all_time_pct != 0 else value

        rows.append({
            "account": account_name,
            "account_type": "401k",
            "ticker": ticker,
            "description": ticker,
            "asset_class": ASSET_CLASS_MAP.get(ticker, "US Equity"),
            "quantity": shares,
            "price": value / shares if shares > 0 else 0,
            "value": value,
            "cost": cost,
            "unrealized_gl": value - cost,
            "unrealized_gl_pct": all_time_pct,
        })
    return pd.DataFrame(rows)


def load_all_holdings() -> pd.DataFrame:
    """Load and normalize all holdings from all account files."""
    frames = []

    # Chase 82-column format files
    chase_files = {
        "ChaseTaxableBrokerage.csv": "Chase Taxable Brokerage",
        "ChaseParametric.csv": "Chase Parametric",
        "ScottsRothIRA_Chase.csv": "Michael's Roth IRA",
        "ScottsTraditionalIRA.csv": "Michael's Traditional IRA",
        # Demo accounts
        "JointBrokerage.csv": "Joint Brokerage",
        "MichaelsRothIRA.csv": "Michael's Roth IRA",
    }
    for filename, account in chase_files.items():
        fp = DATA_DIR / filename
        if fp.exists():
            df = parse_chase_holdings(fp, account)
            if not df.empty:
                frames.append(df)
                print(f"  Loaded {account}: {len(df)} holdings, ${df['value'].sum():,.2f}")

    # SoFi Joint (managed portfolio)
    fp = DATA_DIR / "SOFI_Joint.csv"
    if fp.exists():
        df = parse_sofi_managed(fp, "SoFi Joint")
        if not df.empty:
            frames.append(df)
            print(f"  Loaded SoFi Joint: {len(df)} holdings, ${df['value'].sum():,.2f}")

    # SoFi Self-Directed (brokerage)
    fp = DATA_DIR / "SOFI_SelfDirected.csv"
    if fp.exists():
        df = parse_sofi_brokerage(fp, "SoFi Self-Directed")
        if not df.empty:
            frames.append(df)
            print(f"  Loaded SoFi Self-Directed: {len(df)} holdings, ${df['value'].sum():,.2f}")

    # Roth IRA accounts (marias_roth parser format)
    roth_files = {
        "Marias Roth IRA.csv": "Viviana's Roth IRA",
        "Vivianas Roth IRA.csv": "Viviana's Roth IRA",  # Demo
    }
    for filename, account in roth_files.items():
        fp = DATA_DIR / filename
        if fp.exists():
            df = parse_marias_roth(fp)
            if not df.empty:
                df["account"] = account
                df["account_type"] = ACCOUNT_TYPES.get(account, "Roth IRA")
                frames.append(df)
                print(f"  Loaded {account}: {len(df)} holdings, ${df['value'].sum():,.2f}")

    # 401k accounts (simple Investment/Balance format)
    simple_401k_files = {
        "Maria Carbon Direct 401k.csv": "Viviana 401k",
        "Michaels 401k.csv": "Michael's 401k",  # Demo
    }
    for filename, account in simple_401k_files.items():
        fp = DATA_DIR / filename
        if fp.exists():
            df = parse_401k_simple(fp, account)
            if not df.empty:
                frames.append(df)
                print(f"  Loaded {account}: {len(df)} holdings, ${df['value'].sum():,.2f}")

    # 401k accounts (Vanguard format)
    vanguard_401k_files = {
        "Maria Accrue LevelTen 401k.csv": "Viviana Secondary 401k",
        "Vivianas 401k.csv": "Viviana's 401k",  # Demo
    }
    for filename, account in vanguard_401k_files.items():
        fp = DATA_DIR / filename
        if fp.exists():
            df = parse_401k_vanguard(fp, account)
            if not df.empty:
                frames.append(df)
                print(f"  Loaded {account}: {len(df)} holdings, ${df['value'].sum():,.2f}")

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


def write_header(ws, headers, row=1):
    """Write a styled header row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def generate_report(holdings: pd.DataFrame):
    """Generate the Excel portfolio report."""
    wb = Workbook()
    total_portfolio = holdings["value"].sum()
    total_cost = holdings[holdings["cost"] > 0]["cost"].sum()
    total_gl = holdings["unrealized_gl"].sum()

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: Portfolio Summary
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Portfolio Summary"

    # -- Overview --
    ws.cell(row=1, column=1, value="PORTFOLIO OVERVIEW").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="As of:").font = Font(bold=True)
    ws.cell(row=2, column=2, value="March 9, 2026")

    overview_data = [
        ("Total Portfolio Value", total_portfolio),
        ("Total Cost Basis", total_cost),
        ("Total Unrealized Gain/Loss", total_gl),
        ("Overall Return", total_gl / total_cost if total_cost > 0 else 0),
    ]
    row = 4
    for label, val in overview_data:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=val)
        if "Return" in label:
            cell.number_format = PCT_FMT
        else:
            cell.number_format = MONEY_FMT
            if "Gain" in label:
                cell.font = GAIN_FONT if val >= 0 else LOSS_FONT
        row += 1

    # -- By Account --
    row += 1
    ws.cell(row=row, column=1, value="BY ACCOUNT").font = Font(bold=True, size=12)
    row += 1
    headers = ["Account", "Account Type", "Value", "Cost Basis", "Unrealized G/L", "G/L %", "% of Portfolio"]
    write_header(ws, headers, row)
    row += 1

    acct_summary = holdings.groupby(["account", "account_type"]).agg(
        value=("value", "sum"),
        cost=("cost", "sum"),
        gl=("unrealized_gl", "sum"),
    ).reset_index().sort_values("value", ascending=False)

    for _, r in acct_summary.iterrows():
        ws.cell(row=row, column=1, value=r["account"])
        ws.cell(row=row, column=2, value=r["account_type"])
        ws.cell(row=row, column=3, value=r["value"]).number_format = MONEY_FMT
        ws.cell(row=row, column=4, value=r["cost"]).number_format = MONEY_FMT
        cell_gl = ws.cell(row=row, column=5, value=r["gl"])
        cell_gl.number_format = MONEY_FMT
        cell_gl.font = GAIN_FONT if r["gl"] >= 0 else LOSS_FONT
        gl_pct = r["gl"] / r["cost"] if r["cost"] > 0 else 0
        cell_pct = ws.cell(row=row, column=6, value=gl_pct)
        cell_pct.number_format = PCT_FMT
        cell_pct.font = GAIN_FONT if gl_pct >= 0 else LOSS_FONT
        ws.cell(row=row, column=7, value=r["value"] / total_portfolio).number_format = PCT_FMT
        row += 1

    # Total row
    for col in range(1, 8):
        ws.cell(row=row, column=col).font = TOTAL_FONT
        ws.cell(row=row, column=col).fill = TOTAL_FILL
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=3, value=total_portfolio).number_format = MONEY_FMT
    ws.cell(row=row, column=4, value=total_cost).number_format = MONEY_FMT
    ws.cell(row=row, column=5, value=total_gl).number_format = MONEY_FMT
    ws.cell(row=row, column=6, value=total_gl / total_cost if total_cost > 0 else 0).number_format = PCT_FMT
    ws.cell(row=row, column=7, value=1.0).number_format = PCT_FMT
    row += 2

    # -- By Account Type (Tax Treatment) --
    ws.cell(row=row, column=1, value="BY TAX TREATMENT").font = Font(bold=True, size=12)
    row += 1
    headers = ["Account Type", "Value", "% of Portfolio"]
    write_header(ws, headers, row)
    row += 1

    tax_summary = holdings.groupby("account_type")["value"].sum().sort_values(ascending=False)
    for acct_type, val in tax_summary.items():
        ws.cell(row=row, column=1, value=acct_type)
        ws.cell(row=row, column=2, value=val).number_format = MONEY_FMT
        ws.cell(row=row, column=3, value=val / total_portfolio).number_format = PCT_FMT
        row += 1

    row += 2

    # -- Asset Allocation --
    ws.cell(row=row, column=1, value="ASSET ALLOCATION").font = Font(bold=True, size=12)
    row += 1
    headers = ["Asset Class", "Value", "% of Portfolio"]
    write_header(ws, headers, row)
    row += 1

    asset_order = [
        "US Equity", "International Equity", "Fixed Income",
        "International Fixed Income", "Real Estate", "Commodities (Gold)",
        "Alternatives", "Crypto", "Balanced/Multi-Asset", "Cash", "Other",
    ]
    asset_summary = holdings.groupby("asset_class")["value"].sum().sort_values(ascending=False)
    for ac in asset_order:
        if ac in asset_summary.index:
            val = asset_summary[ac]
            ws.cell(row=row, column=1, value=ac)
            ws.cell(row=row, column=2, value=val).number_format = MONEY_FMT
            ws.cell(row=row, column=3, value=val / total_portfolio).number_format = PCT_FMT
            row += 1
    # Any not in order
    for ac, val in asset_summary.items():
        if ac not in asset_order:
            ws.cell(row=row, column=1, value=ac)
            ws.cell(row=row, column=2, value=val).number_format = MONEY_FMT
            ws.cell(row=row, column=3, value=val / total_portfolio).number_format = PCT_FMT
            row += 1

    row += 2

    # -- Asset Class by Retirement vs Taxable --
    ws.cell(row=row, column=1, value="ASSET CLASS BY RETIREMENT vs TAXABLE").font = Font(bold=True, size=12)
    row += 1

    # Group account types into Retirement and Taxable
    retirement_types = {"401k", "Roth IRA", "Traditional IRA"}
    holdings["tax_bucket"] = holdings["account_type"].apply(
        lambda t: "Retirement" if t in retirement_types else "Taxable"
    )
    buckets = ["Retirement", "Taxable"]

    # Compute bucket totals for percentage columns
    bucket_totals = {b: holdings[holdings["tax_bucket"] == b]["value"].sum() for b in buckets}

    headers = ["Asset Class", "Retirement", "% of Retirement", "Taxable", "% of Taxable", "Total", "% of Portfolio"]
    write_header(ws, headers, row)
    row += 1

    def write_retirement_taxable_row(ws, row, ac_name, holdings_df):
        ws.cell(row=row, column=1, value=ac_name).font = Font(bold=True)
        ac_total = 0
        for bi, bucket in enumerate(buckets):
            val = holdings_df[(holdings_df["asset_class"] == ac_name) & (holdings_df["tax_bucket"] == bucket)]["value"].sum()
            col_val = 2 + bi * 2  # columns: 2 (Retirement $), 4 (Taxable $)
            col_pct = 3 + bi * 2  # columns: 3 (% of Retirement), 5 (% of Taxable)
            ws.cell(row=row, column=col_val, value=val).number_format = MONEY_FMT
            ws.cell(row=row, column=col_pct, value=val / bucket_totals[bucket] if bucket_totals[bucket] else 0).number_format = PCT_FMT
            ac_total += val
        ws.cell(row=row, column=6, value=ac_total).number_format = MONEY_FMT
        ws.cell(row=row, column=7, value=ac_total / total_portfolio).number_format = PCT_FMT

    for ac in asset_order:
        if ac not in asset_summary.index:
            continue
        write_retirement_taxable_row(ws, row, ac, holdings)
        row += 1
    # Any not in order
    for ac, val in asset_summary.items():
        if ac not in asset_order:
            write_retirement_taxable_row(ws, row, ac, holdings)
            row += 1

    # Total row
    for col in range(1, 8):
        ws.cell(row=row, column=col).font = TOTAL_FONT
        ws.cell(row=row, column=col).fill = TOTAL_FILL
    ws.cell(row=row, column=1, value="TOTAL")
    for bi, bucket in enumerate(buckets):
        ws.cell(row=row, column=2 + bi * 2, value=bucket_totals[bucket]).number_format = MONEY_FMT
        ws.cell(row=row, column=3 + bi * 2, value=1.0).number_format = PCT_FMT
    ws.cell(row=row, column=6, value=total_portfolio).number_format = MONEY_FMT
    ws.cell(row=row, column=7, value=1.0).number_format = PCT_FMT

    # Clean up temp column
    holdings.drop(columns=["tax_bucket"], inplace=True)

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 15

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: Top Holdings
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Top Holdings")

    # Consolidate holdings by ticker (across accounts)
    # For non-ticker holdings, use description
    holdings["key"] = holdings.apply(
        lambda r: r["ticker"] if r["ticker"] else r["description"], axis=1
    )
    consolidated = holdings.groupby("key").agg(
        ticker=("ticker", "first"),
        description=("description", "first"),
        asset_class=("asset_class", "first"),
        value=("value", "sum"),
        cost=("cost", "sum"),
        unrealized_gl=("unrealized_gl", "sum"),
        num_accounts=("account", "nunique"),
        accounts=("account", lambda x: ", ".join(sorted(x.unique()))),
    ).reset_index(drop=True).sort_values("value", ascending=False)

    headers = ["#", "Ticker", "Description", "Asset Class", "Value",
               "Cost", "Unrealized G/L", "G/L %", "% of Portfolio", "Account(s)"]
    write_header(ws2, headers)

    for i, (_, r) in enumerate(consolidated.head(50).iterrows(), 1):
        row = i + 1
        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=r["ticker"])
        ws2.cell(row=row, column=3, value=r["description"][:50])
        ws2.cell(row=row, column=4, value=r["asset_class"])
        ws2.cell(row=row, column=5, value=r["value"]).number_format = MONEY_FMT
        ws2.cell(row=row, column=6, value=r["cost"]).number_format = MONEY_FMT
        cell_gl = ws2.cell(row=row, column=7, value=r["unrealized_gl"])
        cell_gl.number_format = MONEY_FMT
        cell_gl.font = GAIN_FONT if r["unrealized_gl"] >= 0 else LOSS_FONT
        gl_pct = r["unrealized_gl"] / r["cost"] if r["cost"] > 0 else 0
        cell_pct = ws2.cell(row=row, column=8, value=gl_pct)
        cell_pct.number_format = PCT_FMT
        cell_pct.font = GAIN_FONT if gl_pct >= 0 else LOSS_FONT
        ws2.cell(row=row, column=9, value=r["value"] / total_portfolio).number_format = PCT_FMT
        ws2.cell(row=row, column=10, value=r["accounts"])

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 15
    ws2.column_dimensions["F"].width = 15
    ws2.column_dimensions["G"].width = 15
    ws2.column_dimensions["H"].width = 10
    ws2.column_dimensions["I"].width = 14
    ws2.column_dimensions["J"].width = 40
    ws2.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: Asset Allocation by Account
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Allocation by Account")

    accounts = holdings.groupby("account")["value"].sum().sort_values(ascending=False).index.tolist()
    asset_classes = [ac for ac in asset_order if ac in holdings["asset_class"].unique()]

    headers = ["Asset Class"] + accounts + ["Total", "% of Portfolio"]
    write_header(ws3, headers)

    for i, ac in enumerate(asset_classes):
        row = i + 2
        ws3.cell(row=row, column=1, value=ac).font = Font(bold=True)
        total_ac = 0
        for j, acct in enumerate(accounts):
            val = holdings[(holdings["asset_class"] == ac) & (holdings["account"] == acct)]["value"].sum()
            ws3.cell(row=row, column=2 + j, value=val).number_format = MONEY_FMT
            total_ac += val
        ws3.cell(row=row, column=2 + len(accounts), value=total_ac).number_format = MONEY_FMT
        ws3.cell(row=row, column=3 + len(accounts), value=total_ac / total_portfolio).number_format = PCT_FMT

    # Total row
    total_row = len(asset_classes) + 2
    for col in range(1, len(headers) + 1):
        ws3.cell(row=total_row, column=col).font = TOTAL_FONT
        ws3.cell(row=total_row, column=col).fill = TOTAL_FILL
    ws3.cell(row=total_row, column=1, value="TOTAL")
    for j, acct in enumerate(accounts):
        val = holdings[holdings["account"] == acct]["value"].sum()
        ws3.cell(row=total_row, column=2 + j, value=val).number_format = MONEY_FMT
    ws3.cell(row=total_row, column=2 + len(accounts), value=total_portfolio).number_format = MONEY_FMT
    ws3.cell(row=total_row, column=3 + len(accounts), value=1.0).number_format = PCT_FMT

    ws3.column_dimensions["A"].width = 25
    for j in range(len(accounts)):
        ws3.column_dimensions[get_column_letter(2 + j)].width = 20
    ws3.column_dimensions[get_column_letter(2 + len(accounts))].width = 15
    ws3.column_dimensions[get_column_letter(3 + len(accounts))].width = 14
    ws3.freeze_panes = "B2"

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 4: All Holdings Detail
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("All Holdings")
    headers = ["Account", "Account Type", "Ticker", "Description", "Asset Class",
               "Quantity", "Price", "Value", "Cost", "Unrealized G/L", "G/L %"]
    write_header(ws4, headers)

    sorted_h = holdings.sort_values(["account", "value"], ascending=[True, False])
    for i, (_, r) in enumerate(sorted_h.iterrows()):
        row = i + 2
        ws4.cell(row=row, column=1, value=r["account"])
        ws4.cell(row=row, column=2, value=r["account_type"])
        ws4.cell(row=row, column=3, value=r["ticker"])
        ws4.cell(row=row, column=4, value=str(r["description"])[:60])
        ws4.cell(row=row, column=5, value=r["asset_class"])
        ws4.cell(row=row, column=6, value=r["quantity"]).number_format = "#,##0.000"
        ws4.cell(row=row, column=7, value=r["price"]).number_format = MONEY_FMT
        ws4.cell(row=row, column=8, value=r["value"]).number_format = MONEY_FMT
        ws4.cell(row=row, column=9, value=r["cost"]).number_format = MONEY_FMT
        cell_gl = ws4.cell(row=row, column=10, value=r["unrealized_gl"])
        cell_gl.number_format = MONEY_FMT
        if r["unrealized_gl"] != 0:
            cell_gl.font = GAIN_FONT if r["unrealized_gl"] >= 0 else LOSS_FONT
        cell_pct = ws4.cell(row=row, column=11, value=r["unrealized_gl_pct"])
        cell_pct.number_format = PCT_FMT
        if r["unrealized_gl_pct"] != 0:
            cell_pct.font = GAIN_FONT if r["unrealized_gl_pct"] >= 0 else LOSS_FONT

    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 45
    ws4.column_dimensions["E"].width = 22
    ws4.column_dimensions["F"].width = 12
    ws4.column_dimensions["G"].width = 12
    ws4.column_dimensions["H"].width = 15
    ws4.column_dimensions["I"].width = 15
    ws4.column_dimensions["J"].width = 15
    ws4.column_dimensions["K"].width = 10
    ws4.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"\nReport saved to: {OUTPUT}")

    # Also save a top-level copy for quick access
    top_level = PORTFOLIO_ROOT / "Portfolio_Report.xlsx"
    wb.save(top_level)
    print(f"Top-level copy saved to: {top_level}")


def get_latest_snapshot_dir() -> Path:
    """Find the most recent snapshot folder, or create today's."""
    from datetime import date
    today = date.today().strftime("%Y-%m-%d")
    snapshot_dir = PORTFOLIO_ROOT / today
    if snapshot_dir.exists():
        return snapshot_dir
    # Find most recent existing snapshot
    if PORTFOLIO_ROOT.exists():
        snapshots = sorted([d for d in PORTFOLIO_ROOT.iterdir() if d.is_dir()])
        if snapshots:
            return snapshots[-1]
    return snapshot_dir


if __name__ == "__main__":
    from datetime import date

    # Use today's snapshot folder, or the latest available
    today = date.today().strftime("%Y-%m-%d")
    snapshot_dir = PORTFOLIO_ROOT / today

    # Look for CSVs in Investments&Balances subfolder or directly in snapshot
    inv_dir = snapshot_dir / "Investments&Balances"
    if inv_dir.exists():
        DATA_DIR = inv_dir
    elif snapshot_dir.exists():
        DATA_DIR = snapshot_dir
    else:
        # Fall back to latest snapshot
        latest = get_latest_snapshot_dir()
        inv_dir = latest / "Investments&Balances"
        DATA_DIR = inv_dir if inv_dir.exists() else latest
        snapshot_dir = latest
        print(f"No snapshot for {today}, using: {snapshot_dir.name}")

    OUTPUT = snapshot_dir / "Portfolio_Report.xlsx"

    print(f"Loading portfolio holdings from {snapshot_dir.name}...\n")
    holdings = load_all_holdings()
    if not holdings.empty:
        total = holdings["value"].sum()
        print(f"\n{'='*60}")
        print(f"Total Portfolio Value: ${total:,.2f}")
        print(f"{'='*60}")
        generate_report(holdings)
