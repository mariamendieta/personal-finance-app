"""
Generate a formatted Excel cash flow report with monthly breakdown
by category and subcategory.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

DATA_FILE = Path(__file__).parent / "CashFlow" / "all_transactions.csv"
OUTPUT_FILE = Path(__file__).parent / "CashFlow" / "CashFlow_Report.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CATEGORY_FONT = Font(name="Calibri", bold=True, size=11)
CATEGORY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBCAT_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", bold=True, size=12)
TOTAL_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
TOTAL_FONT_WHITE = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
INCOME_FONT = Font(name="Calibri", bold=True, size=11, color="006100")
INCOME_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NET_FONT = Font(name="Calibri", bold=True, size=12)
NET_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)
THICK_BORDER = Border(
    top=Side(style="medium", color="2F5496"),
    bottom=Side(style="medium", color="2F5496"),
)
MONEY_FMT = '#,##0.00'
MONEY_FMT_NEG = '#,##0.00;[Red](#,##0.00)'

# Category display order
CATEGORY_ORDER = [
    "Mortgage & Student Loans",
    "Car",
    "Utilities",
    "Childcare",
    "Therapy & Coaching",
    "House & Maintenance",
    "Fitness & Healthcare",
    "Travel",
    "Groceries",
    "Restaurants",
    "Subscriptions",
    "Shopping",
    "Fun & Entertainment",
    "Donations",
    "Pets",
    "Food, Fun & Shopping",
    "Taxes & Tax Fees",
    "Fees & Bank Charges",
]


def generate_report():
    df = pd.read_csv(DATA_FILE)
    df["date"] = pd.to_datetime(df["date"])
    df["month"] = df["date"].dt.to_period("M")
    df["abs_amount"] = df["amount"].abs()

    spending = df[df["flow_type"] == "spending"].copy()
    income = df[df["flow_type"] == "income"].copy()

    months = sorted(spending["month"].unique())
    month_labels = [str(m) for m in months]

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Cash Flow"

    # ── Header row ────────────────────────────────────────────────────────
    headers = ["Category", "Subcategory"] + month_labels + ["Total", "Monthly Avg"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row = 2

    # ── Income section ────────────────────────────────────────────────────
    income_by_month = {}
    for m in months:
        income_by_month[str(m)] = income[income["month"] == m]["amount"].sum()
    income_total = sum(income_by_month.values())

    # Income category header row
    cell = ws.cell(row=row, column=1, value="INCOME")
    cell.font = INCOME_FONT
    cell.fill = INCOME_FILL
    ws.cell(row=row, column=2, value="").fill = INCOME_FILL
    for ci, m in enumerate(month_labels):
        cell = ws.cell(row=row, column=3 + ci, value=income_by_month.get(m, 0))
        cell.number_format = MONEY_FMT
        cell.font = INCOME_FONT
        cell.fill = INCOME_FILL
    cell = ws.cell(row=row, column=3 + len(month_labels), value=income_total)
    cell.number_format = MONEY_FMT
    cell.font = INCOME_FONT
    cell.fill = INCOME_FILL
    cell = ws.cell(row=row, column=4 + len(month_labels), value=income_total / max(len(months), 1))
    cell.number_format = MONEY_FMT
    cell.font = INCOME_FONT
    cell.fill = INCOME_FILL
    row += 1

    # Income subcategory rows
    income_subcats = income.groupby("subcategory")
    for subcat_name in sorted(income_subcats.groups.keys(), key=lambda s: (s == "", s)):
        sub_df = income_subcats.get_group(subcat_name)
        label = subcat_name if subcat_name else "(untagged)"
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value=label).font = Font(italic=True)
        sub_total = 0
        for ci, m in enumerate(month_labels):
            val = sub_df[sub_df["month"] == months[ci]]["amount"].sum() if months[ci] in sub_df["month"].values else 0
            ws.cell(row=row, column=3 + ci, value=val).number_format = MONEY_FMT
            sub_total += val
        ws.cell(row=row, column=3 + len(month_labels), value=sub_total).number_format = MONEY_FMT
        ws.cell(row=row, column=4 + len(month_labels), value=sub_total / max(len(months), 1)).number_format = MONEY_FMT
        row += 1

    row += 1  # blank row

    # ── Spending by category & subcategory ─────────────────────────────────
    spending_start_row = row
    grand_totals = {m: 0 for m in month_labels}
    grand_total = 0

    # Get categories in order, including any not in CATEGORY_ORDER
    present_cats = spending["category"].unique()
    ordered_cats = [c for c in CATEGORY_ORDER if c in present_cats]
    remaining = [c for c in present_cats if c not in ordered_cats]
    ordered_cats.extend(sorted(remaining))

    for cat in ordered_cats:
        cat_data = spending[spending["category"] == cat]
        cat_monthly = {}
        for m in months:
            cat_monthly[str(m)] = cat_data[cat_data["month"] == m]["abs_amount"].sum()
        cat_total = sum(cat_monthly.values())
        grand_total += cat_total
        for m in month_labels:
            grand_totals[m] += cat_monthly.get(m, 0)

        # Category row
        cell = ws.cell(row=row, column=1, value=cat)
        cell.font = CATEGORY_FONT
        cell.fill = CATEGORY_FILL
        ws.cell(row=row, column=2, value="").fill = CATEGORY_FILL
        for ci, m in enumerate(month_labels):
            cell = ws.cell(row=row, column=3 + ci, value=cat_monthly.get(m, 0))
            cell.number_format = MONEY_FMT
            cell.font = CATEGORY_FONT
            cell.fill = CATEGORY_FILL
        cell = ws.cell(row=row, column=3 + len(month_labels), value=cat_total)
        cell.number_format = MONEY_FMT
        cell.font = CATEGORY_FONT
        cell.fill = CATEGORY_FILL
        cell = ws.cell(row=row, column=4 + len(month_labels), value=cat_total / max(len(months), 1))
        cell.number_format = MONEY_FMT
        cell.font = CATEGORY_FONT
        cell.fill = CATEGORY_FILL
        row += 1

        # Subcategory rows
        subcats = cat_data.groupby("subcategory")["abs_amount"].sum().sort_values(ascending=False)
        for subcat, _ in subcats.items():
            sub_data = cat_data[cat_data["subcategory"] == subcat]
            label = subcat if subcat else "(other)"
            ws.cell(row=row, column=1, value="")
            cell = ws.cell(row=row, column=2, value=label)
            cell.font = SUBCAT_FONT
            for ci, m in enumerate(month_labels):
                val = sub_data[sub_data["month"] == months[ci]]["abs_amount"].sum()
                cell = ws.cell(row=row, column=3 + ci, value=val if val > 0 else "")
                cell.number_format = MONEY_FMT
                cell.font = SUBCAT_FONT
            sub_total = sub_data["abs_amount"].sum()
            cell = ws.cell(row=row, column=3 + len(month_labels), value=sub_total)
            cell.number_format = MONEY_FMT
            cell.font = SUBCAT_FONT
            cell = ws.cell(row=row, column=4 + len(month_labels), value=sub_total / max(len(months), 1))
            cell.number_format = MONEY_FMT
            cell.font = SUBCAT_FONT
            # Subtle border
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        row += 1  # gap between categories

    # ── Total spending row ────────────────────────────────────────────────
    cell = ws.cell(row=row, column=1, value="TOTAL SPENDING")
    cell.font = TOTAL_FONT_WHITE
    cell.fill = TOTAL_FILL
    ws.cell(row=row, column=2, value="").fill = TOTAL_FILL
    for ci, m in enumerate(month_labels):
        cell = ws.cell(row=row, column=3 + ci, value=grand_totals[m])
        cell.number_format = MONEY_FMT
        cell.font = TOTAL_FONT_WHITE
        cell.fill = TOTAL_FILL
    cell = ws.cell(row=row, column=3 + len(month_labels), value=grand_total)
    cell.number_format = MONEY_FMT
    cell.font = TOTAL_FONT_WHITE
    cell.fill = TOTAL_FILL
    cell = ws.cell(row=row, column=4 + len(month_labels), value=grand_total / max(len(months), 1))
    cell.number_format = MONEY_FMT
    cell.font = TOTAL_FONT_WHITE
    cell.fill = TOTAL_FILL
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = THICK_BORDER
    row += 1

    # ── Net cash flow row ─────────────────────────────────────────────────
    row += 1
    cell = ws.cell(row=row, column=1, value="NET CASH FLOW")
    cell.font = NET_FONT
    cell.fill = NET_FILL
    ws.cell(row=row, column=2, value="").fill = NET_FILL
    net_total = 0
    for ci, m in enumerate(month_labels):
        net = income_by_month.get(m, 0) - grand_totals[m]
        net_total += net
        cell = ws.cell(row=row, column=3 + ci, value=net)
        cell.number_format = MONEY_FMT_NEG
        cell.font = NET_FONT
        cell.fill = NET_FILL
    cell = ws.cell(row=row, column=3 + len(month_labels), value=net_total)
    cell.number_format = MONEY_FMT_NEG
    cell.font = NET_FONT
    cell.fill = NET_FILL
    cell = ws.cell(row=row, column=4 + len(month_labels), value=net_total / max(len(months), 1))
    cell.number_format = MONEY_FMT_NEG
    cell.font = NET_FONT
    cell.fill = NET_FILL
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = THICK_BORDER

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 32
    for ci in range(len(month_labels)):
        ws.column_dimensions[get_column_letter(3 + ci)].width = 14
    ws.column_dimensions[get_column_letter(3 + len(month_labels))].width = 14
    ws.column_dimensions[get_column_letter(4 + len(month_labels))].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Transactions detail sheet ─────────────────────────────────────────
    ws2 = wb.create_sheet("All Transactions")
    detail_cols = ["date", "description", "amount", "account", "account_type",
                   "flow_type", "category", "subcategory", "original_category", "source_file"]
    for ci, col in enumerate(detail_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for ri, (_, txn) in enumerate(df.sort_values("date").iterrows(), 2):
        for ci, col in enumerate(detail_cols, 1):
            val = txn[col]
            cell = ws2.cell(row=ri, column=ci, value=val)
            if col == "amount":
                cell.number_format = MONEY_FMT_NEG

    # Auto-width for detail sheet
    for ci, col in enumerate(detail_cols, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = max(len(col) + 4, 14)
    ws2.column_dimensions["B"].width = 45
    ws2.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    print(f"Report saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    generate_report()
